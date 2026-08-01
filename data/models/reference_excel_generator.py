"""
excel_generator.py - Gerador de Planilha Excel
================================================
Gera a planilha final no layout do modelo 0001_2026_lctos.xlsx:

Colunas:
  Lançamento | Data | Débito | Crédito | Valor | Histórico Padrão |
  Complemento | CCDB | CCCR | CNPJ

Formatação:
  - Data: DD.MM.AAAA
  - Valor: 2 casas decimais
  - Cabeçalho: negrito, fundo azul escuro
  - Linhas alternadas com cor de fundo para legibilidade

Uso:
    from excel_generator import ExcelGenerator
    gen = ExcelGenerator()
    caminho = gen.gerar("Empresa XYZ", "02", "2026", lancamentos)
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Gera a planilha Excel de lançamentos contábeis."""

    # Cores da planilha (padrão dark blue)
    COR_HEADER_BG  = "1E2A4A"   # Azul escuro cabeçalho
    COR_HEADER_FG  = "FFFFFF"   # Branco texto cabeçalho
    COR_LINHA_PAR  = "F0F4FF"   # Azul bem claro linha par
    COR_LINHA_IMPAR = "FFFFFF"  # Branco linha ímpar
    COR_BORDA      = "C5CCE5"   # Cinza borda

    def gerar(
        self,
        empresa: str,
        mes: str,
        ano: str,
        lancamentos: list[dict],
    ) -> Path:
        """
        Gera a planilha Excel com os lançamentos aprovados.

        Returns:
            Path do arquivo .xlsx gerado.
        """
        nome_mes = {
            "01": "Janeiro", "02": "Fevereiro", "03": "Março",
            "04": "Abril", "05": "Maio", "06": "Junho",
            "07": "Julho", "08": "Agosto", "09": "Setembro",
            "10": "Outubro", "11": "Novembro", "12": "Dezembro",
        }.get(mes, mes)

        empresa_slug = empresa.replace(" ", "_").replace("/", "-")
        nome_arquivo = f"Lancamentos_{empresa_slug}_{mes}{ano}.xlsx"
        caminho = config.OUTPUT_DIR / nome_arquivo

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lançamentos"

        # --- Título da empresa no topo ---
        ws.merge_cells("A1:J1")
        celula_titulo = ws["A1"]
        celula_titulo.value = f"{empresa} — {nome_mes}/{ano}"
        celula_titulo.font = Font(
            name="Calibri", bold=True, size=14, color="1E2A4A"
        )
        celula_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:J2")
        celula_data = ws["A2"]
        celula_data.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        celula_data.font = Font(name="Calibri", size=10, color="888888")
        celula_data.alignment = Alignment(horizontal="right")
        ws.row_dimensions[2].height = 16

        # --- Linha em branco separadora ---
        ws.row_dimensions[3].height = 8

        # --- Cabeçalho das colunas (linha 4) ---
        colunas = config.EXCEL_COLUMNS
        fill_header = PatternFill("solid", fgColor=self.COR_HEADER_BG)
        borda = self._borda_simples()

        for col_idx, nome_col in enumerate(colunas, start=1):
            cell = ws.cell(row=4, column=col_idx, value=nome_col)
            cell.font = Font(
                name="Calibri", bold=True, size=11, color=self.COR_HEADER_FG
            )
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borda

        ws.row_dimensions[4].height = 22

        # --- Dados dos lançamentos (a partir da linha 5) ---
        for i, lanc in enumerate(lancamentos):
            linha = 5 + i
            num_lancamento = i + 1
            fill_linha = PatternFill(
                "solid",
                fgColor=self.COR_LINHA_PAR if i % 2 == 0 else self.COR_LINHA_IMPAR
            )

            # Formata a data para DD.MM.AAAA
            data_str = self._formatar_data(lanc.get("data_lancamento", ""))

            valores_linha = [
                num_lancamento,
                data_str,
                lanc.get("conta_debito", ""),
                lanc.get("conta_credito", ""),
                lanc.get("valor", 0),
                lanc.get("historico", ""),
                lanc.get("complemento", ""),
                lanc.get("ccdb", ""),
                lanc.get("cccr", ""),
                lanc.get("cnpj", ""),
            ]

            for col_idx, valor in enumerate(valores_linha, start=1):
                cell = ws.cell(row=linha, column=col_idx, value=valor)
                cell.fill = fill_linha
                cell.border = borda
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Coluna Valor (índice 5): formato número com 2 casas
                if col_idx == 5:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # Coluna Lançamento (índice 1): centralizado
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Histórico (índice 6): esquerda com wrap
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            ws.row_dimensions[linha].height = 18

        # --- Ajuste de largura das colunas ---
        larguras = {
            1: 12,   # Lançamento
            2: 14,   # Data
            3: 14,   # Débito
            4: 14,   # Crédito
            5: 14,   # Valor
            6: 55,   # Histórico Padrão (mais largo)
            7: 30,   # Complemento
            8: 12,   # CCDB
            9: 12,   # CCCR
            10: 20,  # CNPJ
        }
        for col_idx, largura in larguras.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

        # --- Congela o painel no cabeçalho ---
        ws.freeze_panes = "A5"

        # --- Filtro automático ---
        ws.auto_filter.ref = f"A4:{get_column_letter(len(colunas))}4"

        wb.save(caminho)
        logger.info("Excel gerado: %s (%d lançamentos)", caminho, len(lancamentos))
        return caminho

    def _borda_simples(self) -> Border:
        lado = Side(style="thin", color=self.COR_BORDA)
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    @staticmethod
    def _formatar_data(data_str: str) -> str:
        """
        Garante que a data esteja no formato DD.MM.AAAA.
        Aceita entradas como: DD/MM/AAAA, AAAA-MM-DD, DD.MM.AAAA
        """
        if not data_str:
            return ""
        # Já está no formato correto
        if len(data_str) == 10 and data_str[2] == "." and data_str[5] == ".":
            return data_str
        # Tenta vários formatos comuns
        formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d"]
        for fmt in formatos:
            try:
                return datetime.strptime(data_str, fmt).strftime("%d.%m.%Y")
            except ValueError:
                continue
        return data_str  # Retorna como está se não conseguir parsear
