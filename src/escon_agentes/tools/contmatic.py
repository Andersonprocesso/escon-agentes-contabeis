"""Exportação Contmatic — layout Escon_Lancamento + motor Contabilizador."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from escon_agentes.tools.plano_contas import banco_codigo, contas, historicos

# Layout oficial do modelo 0001_2026_lctos.xlsx (Escon_Lancamento)
EXCEL_COLUMNS = [
    "Lançamento",
    "Data",
    "Débito",
    "Crédito",
    "Valor",
    "Histórico Padrão",
    "Complemento",
    "CCDB",
    "CCCR",
    "CNPJ",
]

HISTORICOS = historicos()

COR_HEADER_BG = "1E2A4A"
COR_HEADER_FG = "FFFFFF"
COR_LINHA_PAR = "F0F4FF"
COR_LINHA_IMPAR = "FFFFFF"
COR_BORDA = "C5CCE5"


def format_valor_br(valor: Any) -> str:
    if isinstance(valor, (int, float)):
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    s = str(valor or "0").strip()
    if not s:
        return "0,00"
    if "." in s and "," not in s:
        try:
            return format_valor_br(float(s))
        except ValueError:
            return s.replace(".", ",")
    return s


def format_data_contmatic(data: Any) -> str:
    """Contmatic / modelo Escon_Lancamento: DD.MM.AAAA"""
    if data is None or data == "":
        return datetime.now().strftime("%d.%m.%Y")
    if isinstance(data, datetime):
        return data.strftime("%d.%m.%Y")
    if isinstance(data, date):
        return data.strftime("%d.%m.%Y")
    s = str(data).strip()
    if len(s) == 10 and s[2] == "." and s[5] == ".":
        return s
    for fmt in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return s


def write_lancamentos(
    rows: list[dict[str, Any]],
    out_path: Path,
    *,
    empresa: str | None = None,
    competencia: str | None = None,
) -> Path:
    """
    Gera Excel no layout do modelo Contmatic Escon (10 colunas).
    Compatível com importação de planilha no Contmatic.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    borda = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )
    fill_header = PatternFill("solid", fgColor=COR_HEADER_BG)

    # Título
    ws.merge_cells("A1:J1")
    titulo = empresa or "Escon Soluções Contábeis"
    if competencia:
        titulo = f"{titulo} — {competencia}"
    c1 = ws["A1"]
    c1.value = titulo
    c1.font = Font(name="Calibri", bold=True, size=14, color="1E2A4A")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    c2 = ws["A2"]
    c2.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Plano Contmatic Escon"
    c2.font = Font(name="Calibri", size=10, color="888888")
    c2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[3].height = 8

    header_row = 4
    for col_idx, nome in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=nome)
        cell.font = Font(name="Calibri", bold=True, size=11, color=COR_HEADER_FG)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[header_row].height = 22

    for i, row in enumerate(rows):
        r = header_row + 1 + i
        fill = PatternFill(
            "solid",
            fgColor=COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR,
        )
        valor = row.get("valor", 0)
        # numérico se possível
        valor_cell: Any = valor
        if isinstance(valor, str):
            try:
                valor_cell = float(
                    valor.replace("R$", "").replace(".", "").replace(",", ".").strip()
                )
            except ValueError:
                valor_cell = valor
        elif isinstance(valor, (int, float)):
            valor_cell = float(valor)

        hist = row.get("historico", 26)
        # Contmatic aceita código numérico de histórico padrão
        valores = [
            row.get("lancamento", i + 1),
            format_data_contmatic(row.get("data")),
            row.get("debito", ""),
            row.get("credito", ""),
            valor_cell,
            hist,
            row.get("complemento", ""),
            row.get("ccdb", row.get("CCDB", "")),
            row.get("cccr", row.get("CCCR", "")),
            row.get("cnpj", row.get("CNPJ", "")),
        ]
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.fill = fill
            cell.border = borda
            cell.font = Font(name="Calibri", size=11)
            if col_idx == 5 and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

    larguras = {1: 12, 2: 14, 3: 14, 4: 14, 5: 14, 6: 16, 7: 50, 8: 12, 9: 12, 10: 20}
    for col_idx, largura in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(EXCEL_COLUMNS))}4"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def rows_from_engine_lancamentos(lancamentos: list[Any]) -> list[dict]:
    rows: list[dict] = []
    for i, lct in enumerate(lancamentos, start=1):
        data = lct.data
        rows.append(
            {
                "lancamento": i,
                "data": data,
                "debito": lct.deb,
                "credito": lct.cred,
                "valor": lct.valor,
                "historico": lct.hist,
                "complemento": lct.comp,
                "ccdb": "",
                "cccr": "",
                "cnpj": "",
            }
        )
    return rows


def processar_pasta_contmatic(
    pasta: Path,
    out_xlsx: Path,
    *,
    banco: str = "itau",
    empresa: str | None = None,
    competencia: str | None = None,
) -> dict[str, Any]:
    from escon_agentes.tools import contabilizador_engine as eng

    eng.C = contas()  # type: ignore[attr-defined]
    banco_cod = banco_codigo(banco)
    pasta = Path(pasta)
    if not pasta.exists():
        return {"success": False, "summary": f"Pasta não encontrada: {pasta}", "lancamentos": 0}

    todos: list = []
    detalhes: list[dict] = []

    for f in sorted(pasta.rglob("*")):
        if not f.is_file():
            continue
        suf = f.suffix.lower()
        try:
            if suf == ".xml":
                lcts = eng.ler_xml(str(f))
            elif suf == ".ofx":
                lcts = eng.ler_ofx(str(f), banco_cod)
            elif suf == ".pdf":
                lcts = eng.ler_pdf(str(f), banco_cod)
            else:
                continue
            todos.extend(lcts)
            detalhes.append({"arquivo": f.name, "tipo": suf, "lancamentos": len(lcts)})
        except Exception as e:  # noqa: BLE001
            detalhes.append({"arquivo": f.name, "tipo": suf, "erro": str(e), "lancamentos": 0})

    if not todos:
        return {
            "success": False,
            "summary": f"Nenhum lançamento gerado em {pasta}",
            "lancamentos": 0,
            "detalhes": detalhes,
            "banco": banco,
            "banco_codigo": banco_cod,
        }

    todos.sort(key=lambda x: x.data)
    rows = rows_from_engine_lancamentos(todos)
    write_lancamentos(rows, out_xlsx, empresa=empresa, competencia=competencia)

    return {
        "success": True,
        "summary": f"{len(todos)} lançamento(s) Contmatic → {out_xlsx}",
        "lancamentos": len(todos),
        "arquivo": str(out_xlsx),
        "detalhes": detalhes,
        "banco": banco,
        "banco_codigo": banco_cod,
        "layout": "Escon_Lancamento / 0001_2026_lctos.xlsx (10 colunas)",
        "plano": "data/models/PlContas.TXT + config/plano_contas.yaml",
    }


def rows_from_xml_summary(
    docs: list[dict[str, Any]],
    conta_receita: str | int | None = None,
    conta_clientes: str | int | None = None,
) -> list[dict]:
    c = contas()
    receita = int(conta_receita or c["receita_servicos"])
    clientes = int(conta_clientes or c["duplicatas_receber"])
    rows: list[dict] = []
    for d in docs:
        if d.get("tipo") not in {"nfe", "nfce", "nfse"}:
            continue
        data = d.get("data_emissao") or ""
        if "T" in str(data):
            data = str(data)[:10]
            parts = data.split("-")
            if len(parts) == 3:
                data = f"{parts[2]}/{parts[1]}/{parts[0]}"
        rows.append(
            {
                "data": data or datetime.now().strftime("%d/%m/%Y"),
                "debito": clientes,
                "credito": receita,
                "valor": d.get("valor_total") or 0,
                "historico": 9,
                "complemento": (
                    f"NF {d.get('numero') or ''} - "
                    f"{d.get('dest_nome') or d.get('emit_nome') or ''}"
                ).strip(),
                "cnpj": d.get("dest_cnpj") or d.get("emit_cnpj") or "",
            }
        )
    return rows


def rows_from_documents(items: list[dict[str, Any]], banco: str = "itau") -> list[dict]:
    c = contas()
    b = banco_codigo(banco)
    doc_map: dict[str, tuple[int, int, int, str]] = {
        "das": (c["desp_simples"], c["simples_nacional"], 7, "DAS Simples"),
        "darf": (c["inss_pagar"], c["simples_nacional"], 15, "DARF"),
        "prolabore": (c["desp_prolabore"], c["prolabore_pagar"], 1, "Pro-labore"),
        "folha": (c["desp_salarios"], c["salarios_pagar"], 3, "Folha"),
        "fgts": (c["fgts_pagar"], b, 19, "FGTS"),
        "gps": (c["inss_pagar"], b, 15, "GPS/INSS"),
        "nf": (c["duplicatas_receber"], c["receita_servicos"], 9, "NF PDF"),
        "boleto": (c["desp_material_escr"], c["fornecedores"], 26, "Boleto"),
        "desconhecido": (c["desp_material_escr"], c["fornecedores"], 26, "Doc"),
    }
    rows: list[dict] = []
    for it in items:
        dtype = it.get("doc_type") or "desconhecido"
        if dtype == "sem_texto":
            continue
        debito, credito, hist, prefix = doc_map.get(dtype, doc_map["desconhecido"])
        valor = it.get("valor")
        if not valor:
            continue
        rows.append(
            {
                "data": it.get("data") or datetime.now().strftime("%d/%m/%Y"),
                "debito": debito,
                "credito": credito,
                "valor": valor,
                "historico": hist,
                "complemento": f"{prefix} — {it.get('descricao') or Path(it.get('path', '')).name}",
                "cnpj": it.get("cnpj") or "",
            }
        )
    return rows
