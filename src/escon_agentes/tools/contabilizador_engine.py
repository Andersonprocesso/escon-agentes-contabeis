# Motor Contabilizador Escon — espelho da skill contabilizador (códigos Contmatic reais).
# Fonte: ~/.claude/skills/user/contabilizador/contabilizador.py
# Não editar regras de conta sem atualizar config/plano_contas.yaml e a skill.
"""
Contabilizador Automático - Simples Nacional / MEI
Gera planilha Excel de lançamentos contábeis a partir de:
  - XML  → NF-e / NFS-e
  - OFX  → Extrato bancário
  - PDF  → Boletos, DAS, DARF, Folha de pagamento, Recibos

Layout: Lançamento | Data | Débito | Crédito | Valor | Histórico Padrão | Complemento
"""

import os, re, xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import argparse

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

# ─────────────────────────────────────────────
# PLANO DE CONTAS
# ─────────────────────────────────────────────
C = {
    "caixa":1111101,"banco_itau":1112201,"banco_bradesco":1112202,
    "banco_cef":1112203,"banco_inter":1112204,"banco_bb":1112205,
    "banco_santander":1112206,"banco_sicoob":1112207,"banco_nubank":1112211,
    "banco_mercadopago":1112208,
    "duplicatas_receber":1121101,"inss_retido_fonte":1131910,
    "fornecedores":2111101,"simples_nacional":2131101,
    "inss_pagar":2131201,"fgts_pagar":2131202,
    "salarios_pagar":2141101,"prolabore_pagar":2141202,
    "receita_servicos":4111201,"receita_produtos":4111103,
    "desp_prolabore":4121101,"desp_inss_prolabore":4121201,
    "desp_salarios":4121301,"desp_fgts":4121401,"desp_simples":4121501,
    "desp_aluguel":4122101,"desp_energia":4122201,"desp_agua":4122202,
    "desp_telefone":4122301,"desp_honorario":4122401,
    "desp_material_escr":4122501,"desp_combustivel":4122601,
    "desp_manutencao":4122701,"desp_publicidade":4122801,
    "desp_vale_transp":4122901,"desp_alimentacao":4123001,
    "desp_depreciacao":4123101,
}

BANCOS_MAP = {
    "itau":"banco_itau","bradesco":"banco_bradesco","cef":"banco_cef",
    "inter":"banco_inter","bb":"banco_bb","santander":"banco_santander",
    "sicoob":"banco_sicoob","nubank":"banco_nubank","mercadopago":"banco_mercadopago",
}

# ─────────────────────────────────────────────
# CLASSE DE LANÇAMENTO
# ─────────────────────────────────────────────
class L:
    def __init__(self, data, deb, cred, valor, hist, comp=""):
        if isinstance(data, (date, datetime)):
            self.data = data if isinstance(data, date) else data.date()
        else:
            self.data = date.today()
            for fmt in ("%d/%m/%Y","%Y-%m-%d","%d%m%Y"):
                try:
                    self.data = datetime.strptime(str(data)[:10], fmt).date(); break
                except: pass
        self.deb=deb; self.cred=cred
        self.valor=round(float(str(valor).replace(".","").replace(",",".")),2)
        self.hist=hist; self.comp=str(comp or "")[:100]

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def _v(txt):
    """Extrai o principal valor monetário do texto."""
    for p in [r"R\$\s*([\d.]+,\d{2})",r"VALOR\D{0,10}([\d.]+,\d{2})",
              r"TOTAL\D{0,10}([\d.]+,\d{2})",r"\b(\d{1,3}(?:\.\d{3})+,\d{2})\b",
              r"\b(\d+,\d{2})\b"]:
        m=re.search(p,txt,re.I)
        if m:
            try:
                v=float(m.group(1).replace(".","").replace(",",".")); 
                if v>0: return v
            except: pass
    return 0.0

def _d(txt):
    """Extrai data do texto."""
    for p in [r"VENCIMENTO\D{0,5}(\d{2}/\d{2}/\d{4})",r"(\d{2}/\d{2}/\d{4})",
              r"(\d{4}-\d{2}-\d{2})",r"COMPETÊNCIA\D{0,5}(\d{2}/\d{4})"]:
        m=re.search(p,txt,re.I)
        if m:
            s=m.group(1)
            for fmt in ("%d/%m/%Y","%Y-%m-%d"):
                try: return datetime.strptime(s,fmt).date()
                except: pass
            try:
                ms,ano=s.split("/"); return date(int(ano),int(ms),1)
            except: pass
    return date.today()

def _tem(txt, palavras): return any(p in txt for p in palavras)

# ─────────────────────────────────────────────
# XML (NF-e)
# ─────────────────────────────────────────────
def ler_xml(arq):
    lcts=[]
    tree=ET.parse(arq); root=tree.getroot()
    ns=root.tag.split("}")[0]+"}" if root.tag.startswith("{") else ""
    def ft(el,path,d=""):
        n=el.find(f".//{ns}"+f"//{ns}".join(path.split("/"))); return n.text if n is not None else d
    inf=root.find(f".//{ns}infNFe")
    if inf is None: print(f"  [X] infNFe não encontrado em {arq}"); return []
    dh=ft(inf,"ide/dhEmi") or ft(inf,"ide/dEmi")
    try: dt=datetime.strptime(dh[:10],"%Y-%m-%d").date()
    except: dt=date.today()
    num=ft(inf,"ide/nNF"); tp=ft(inf,"ide/tpNF")
    dest=ft(inf,"dest/xNome"); emit=ft(inf,"emit/xNome")
    def fv(s):
        try: return float(str(s).replace(",","."))
        except: return 0.0
    vnf=fv(ft(inf,"total/ICMSTot/vNF"))
    viss=fv(ft(inf,"total/ISSQNtot/vISSRet"))
    vinss=fv(ft(inf,"total/ISSQNtot/vRetPIS"))
    if tp=="1":
        cb=f"NF {num} - {dest}"
        lcts.append(L(dt,C["duplicatas_receber"],C["receita_servicos"],vnf,9,cb))
        if viss>0: lcts.append(L(dt,C["receita_servicos"],C["receita_servicos"],viss,26,"Valor Referente ISS Retido Na Fonte"))
        if vinss>0: lcts.append(L(dt,C["inss_retido_fonte"],C["receita_servicos"],vinss,34,""))
        liq=vnf-viss-vinss
        if liq>0: lcts.append(L(dt,C["caixa"],C["duplicatas_receber"],liq,26,f"NFS {num}"))
    else:
        lcts.append(L(dt,C["desp_material_escr"],C["fornecedores"],vnf,25,f"NF {num} - {emit}"))
    return lcts

# ─────────────────────────────────────────────
# OFX (Extrato)
# ─────────────────────────────────────────────
def ler_ofx(arq, banco=None):
    banco=banco or C["banco_itau"]; lcts=[]
    txt=open(arq,"r",encoding="latin-1",errors="ignore").read()
    for trn in re.findall(r"<STMTTRN>(.*?)</STMTTRN>",txt,re.DOTALL):
        def g(t): m=re.search(rf"<{t}>(.*?)(?:<|\n)",trn); return m.group(1).strip() if m else ""
        tp=g("TRNTYPE"); dp=g("DTPOSTED"); amt=g("TRNAMT"); memo=g("MEMO") or g("NAME") or ""
        try: dt=datetime.strptime(dp[:8],"%Y%m%d").date()
        except: dt=date.today()
        try: v=abs(float(amt.replace(",",".")))
        except: continue
        h,deb,cred=_clf_ofx(memo.upper(),tp,banco)
        lcts.append(L(dt,deb,cred,v,h,memo[:80]))
    return lcts

def _clf_ofx(mu,tp,banco):
    if tp in ("CREDIT","DEP","INT"): return 26,banco,C["duplicatas_receber"]
    for kw,h,deb,cred in [
        (["PROLABORE","PRÓ-LABORE"],36,C["prolabore_pagar"],banco),
        (["SALARIO","SALÁRIO"],35,C["salarios_pagar"],banco),
        (["FGTS"],19,C["fgts_pagar"],banco),
        (["INSS","PREVIDENCIA"],15,C["inss_pagar"],banco),
        (["SIMPLES","DAS ","PGDAS"],38,C["simples_nacional"],banco),
        (["ALUGUEL"],40,C["desp_aluguel"],banco),
        (["ENERGIA","ENEL","CPFL","CEMIG","LIGHT"],11,C["desp_energia"],banco),
        (["SABESP","SANEPAR","COPASA"],10,C["desp_agua"],banco),
        (["TELEFONE","INTERNET","CLARO","VIVO","TIM "],29,C["desp_telefone"],banco),
        (["HONORARIO","CONTABILIDADE"],37,C["desp_honorario"],banco),
        (["COMBUSTIVEL","POSTO","GASOLINA"],12,C["desp_combustivel"],banco),
        (["VALE TRANSP"],8,C["desp_vale_transp"],banco),
        (["ALIMENTACAO","IFOOD","TICKET","SODEXO"],14,C["desp_alimentacao"],banco),
    ]:
        if any(p in mu for p in kw): return h,deb,cred
    return 26,C["fornecedores"],banco

# ─────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────
def ler_pdf(arq, banco=None):
    if not PDF_OK: print(f"  [X] pdfplumber não instalado. Instale: pip install pdfplumber"); return []
    banco=banco or C["banco_itau"]
    try:
        with pdfplumber.open(arq) as pdf:
            txt="\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception as e:
        print(f"  [X] Erro ao ler PDF {Path(arq).name}: {e}"); return []

    tu=txt.upper()
    nome=Path(arq).stem.upper()

    if _tem(tu,["DOCUMENTO DE ARRECADAÇÃO DO SIMPLES","PGDAS","SIMPLES NACIONAL","DAS "]):
        return _pdf_das(txt,tu,banco)
    if _tem(tu,["DARF","DOCUMENTO DE ARRECADAÇÃO DE RECEITAS"]):
        return _pdf_darf(txt,tu,banco)
    if _tem(tu,["PRÓ-LABORE","PRO-LABORE","PROLABORE","PRÓ LABORE"]):
        return _pdf_prolabore(txt,tu,banco)
    if _tem(tu,["FOLHA DE PAGAMENTO","HOLERITE","CONTRACHEQUE","RECIBO DE PAGAMENTO"]):
        return _pdf_folha(txt,tu,banco)
    if _tem(tu,["NOTA FISCAL","NF-E","NFS-E","CHAVE DE ACESSO"]):
        return _pdf_nf(txt,tu,banco)
    if _tem(tu,["FGTS","SEFIP","GUIA DE RECOLHIMENTO DO FGTS"]):
        return _pdf_fgts(txt,tu,banco)
    if _tem(tu,["GPS","GUIA DA PREVIDÊNCIA SOCIAL"]):
        return _pdf_inss(txt,tu,banco)
    if _tem(tu,["BOLETO","LINHA DIGITÁVEL","CÓDIGO DE BARRAS","VENCIMENTO"]):
        return _pdf_boleto(txt,tu,banco,nome)

    print(f"  [?] Tipo não reconhecido: {Path(arq).name}")
    print(f"      Trecho: {txt[:150]}")
    return []

def _pdf_das(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: print("  [!] DAS: valor não encontrado"); return []
    m=re.search(r"COMPETÊNCIA\D{0,5}(\d{2}/\d{4})",txt,re.I)
    comp=m.group(1) if m else ""
    return [L(dt,C["desp_simples"],C["simples_nacional"],v,7,f"DAS {comp}"),
            L(dt,C["simples_nacional"],banco,v,38,f"DAS {comp}")]

def _pdf_darf(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: return []
    return [L(dt,C["inss_pagar"],C["simples_nacional"],v,15,"DARF"),
            L(dt,C["simples_nacional"],banco,v,15,"Pgto DARF")]

def _pdf_prolabore(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: print("  [!] Pró-labore: valor não encontrado"); return []
    m=re.search(r"(?:SÓCIO|SOCIO|NOME)[:\s]+([A-Z\s]{5,40})",tu)
    nome=m.group(1).strip()[:40] if m else ""
    inss=round(v*0.11,2); liq=round(v-inss,2)
    return [L(dt,C["desp_prolabore"],C["prolabore_pagar"],v,1,nome),
            L(dt,C["desp_inss_prolabore"],C["inss_pagar"],inss,2,nome),
            L(dt,C["prolabore_pagar"],banco,liq,36,nome),
            L(dt,C["inss_pagar"],banco,inss,15,nome)]

def _pdf_folha(txt,tu,banco):
    dt=_d(txt); linhas=txt.split("\n")
    for i,linha in enumerate(linhas):
        lu=linha.upper()
        if any(k in lu for k in ["SALÁRIO BRUTO","TOTAL BRUTO","VENCIMENTOS"]):
            v=_v(lu)
            if v>0:
                inss=fgts=0.0
                for j in range(i+1,min(i+10,len(linhas))):
                    lj=linhas[j].upper()
                    if "INSS" in lj: inss=_v(lj) or inss
                    if "FGTS" in lj: fgts=_v(lj) or fgts
                inss=inss or round(v*0.075,2); fgts=fgts or round(v*0.08,2)
                liq=round(v-inss,2)
                nome="".join(linhas[max(0,i-3):i])[:40].strip()
                return [L(dt,C["desp_salarios"],C["salarios_pagar"],v,3,nome),
                        L(dt,C["desp_fgts"],C["fgts_pagar"],fgts,5,nome),
                        L(dt,C["desp_salarios"],C["inss_pagar"],inss,6,nome),
                        L(dt,C["salarios_pagar"],banco,liq,35,nome),
                        L(dt,C["fgts_pagar"],banco,fgts,19,nome),
                        L(dt,C["inss_pagar"],banco,inss,15,nome)]
    v=_v(tu)
    if not v: return []
    inss=round(v*0.075,2); fgts=round(v*0.08,2); liq=round(v-inss,2)
    return [L(dt,C["desp_salarios"],C["salarios_pagar"],v,3,""),
            L(dt,C["desp_fgts"],C["fgts_pagar"],fgts,5,""),
            L(dt,C["salarios_pagar"],banco,liq,35,"")]

def _pdf_nf(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: return []
    m=re.search(r"(?:NF|NOTA FISCAL)\D{0,5}(\d+)",tu)
    num=m.group(1) if m else ""
    eh_serv="SERVIÇOS" in tu or "PRESTAÇÃO" in tu
    cr=C["receita_servicos"] if eh_serv else C["receita_produtos"]
    return [L(dt,C["duplicatas_receber"],cr,v,9,f"NF {num}"),
            L(dt,C["caixa"],C["duplicatas_receber"],v,26,f"NFS {num}")]

def _pdf_boleto(txt,tu,banco,nome_arq):
    v=_v(tu); dt=_d(txt)
    if not v: return []
    h,deb=_clf_boleto(tu,nome_arq)
    return [L(dt,deb,C["fornecedores"],v,h,""),
            L(dt,C["fornecedores"],banco,v,h,"")]

def _clf_boleto(tu,nome):
    for kw,h,deb in [
        (["ALUGUEL","LOCAÇÃO","LOCACAO"],40,C["desp_aluguel"]),
        (["ENERGIA","ENEL","CPFL","CEMIG","LIGHT"],11,C["desp_energia"]),
        (["SABESP","SANEPAR","COPASA","SANEAMENTO"],10,C["desp_agua"]),
        (["TELEFONE","INTERNET","CLARO","VIVO","TIM","NET "],29,C["desp_telefone"]),
        (["CONTABILIDADE","HONORÁRIO","HONORARIO"],37,C["desp_honorario"]),
        (["COMBUSTÍVEL","GASOLINA","POSTO"],12,C["desp_combustivel"]),
        (["MATERIAL","PAPELARIA"],24,C["desp_material_escr"]),
        (["PUBLICIDADE","MARKETING"],23,C["desp_publicidade"]),
        (["IPTU","IMPOSTO PREDIAL"],21,C["desp_material_escr"]),
    ]:
        if any(p in tu or p in nome for p in kw): return h,deb
    return 26,C["desp_material_escr"]

def _pdf_fgts(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: return []
    return [L(dt,C["fgts_pagar"],banco,v,19,"FGTS")]

def _pdf_inss(txt,tu,banco):
    v=_v(tu); dt=_d(txt)
    if not v: return []
    return [L(dt,C["inss_pagar"],banco,v,15,"INSS GPS")]

# ─────────────────────────────────────────────
# LANÇAMENTOS MANUAIS
# ─────────────────────────────────────────────
def lancar_prolabore(data,valor,nome="",inss_pct=0.11,banco=None):
    banco=banco or C["banco_itau"]
    inss=round(valor*inss_pct,2); liq=round(valor-inss,2)
    return [L(data,C["desp_prolabore"],C["prolabore_pagar"],valor,1,nome),
            L(data,C["desp_inss_prolabore"],C["inss_pagar"],inss,2,nome),
            L(data,C["prolabore_pagar"],banco,liq,36,nome),
            L(data,C["inss_pagar"],banco,inss,15,nome)]

def lancar_salario(data,nome,bruto,inss,fgts,banco=None):
    banco=banco or C["banco_itau"]; liq=round(bruto-inss,2)
    return [L(data,C["desp_salarios"],C["salarios_pagar"],bruto,3,nome),
            L(data,C["desp_fgts"],C["fgts_pagar"],fgts,5,nome),
            L(data,C["desp_salarios"],C["inss_pagar"],inss,6,nome),
            L(data,C["salarios_pagar"],banco,liq,35,nome),
            L(data,C["fgts_pagar"],banco,fgts,19,nome),
            L(data,C["inss_pagar"],banco,inss,15,nome)]

def lancar_das(data,valor,comp="",banco=None):
    banco=banco or C["banco_itau"]; tag=f"DAS {comp}".strip()
    return [L(data,C["desp_simples"],C["simples_nacional"],valor,7,tag),
            L(data,C["simples_nacional"],banco,valor,38,tag)]

# ─────────────────────────────────────────────
# GERAR EXCEL
# ─────────────────────────────────────────────
def gerar_excel(lancamentos, saida, num_ini=1):
    wb=Workbook(); ws=wb.active; ws.title="Planilha1"
    hf=Font(name="Calibri",bold=True,color="FFFFFF",size=11)
    hfill=PatternFill("solid",fgColor="1F4E79")
    ha=Alignment(horizontal="center",vertical="center")
    brd=Border(left=Side(style="thin"),right=Side(style="thin"),
               top=Side(style="thin"),bottom=Side(style="thin"))
    cols=["Lançamento","Data","Débito","Crédito","Valor","Histórico Padrão","Complemento"]
    lrgs=[12,12,12,12,14,16,50]
    for ci,(nm,lg) in enumerate(zip(cols,lrgs),1):
        cl=ws.cell(row=1,column=ci,value=nm)
        cl.font=hf; cl.fill=hfill; cl.alignment=ha; cl.border=brd
        ws.column_dimensions[cl.column_letter].width=lg
    ws.row_dimensions[1].height=20
    fp=PatternFill("solid",fgColor="DEEAF1"); fi=PatternFill("solid",fgColor="FFFFFF")
    nf=Font(name="Calibri",size=10)
    for i,lct in enumerate(lancamentos):
        row=i+2; fill=fp if i%2==0 else fi
        vf=f"{lct.valor:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        dados=[num_ini+i,lct.data.strftime("%d/%m/%Y"),lct.deb,lct.cred,vf,lct.hist,lct.comp]
        for ci,val in enumerate(dados,1):
            cl=ws.cell(row=row,column=ci,value=val)
            cl.font=nf; cl.fill=fill; cl.border=brd
            cl.alignment=Alignment(
                horizontal="center" if ci==2 else ("right" if ci in(1,3,4,6) else "left"))
    wb.save(saida)
    print(f"\n✅ Planilha gerada: {saida}  ({len(lancamentos)} lançamentos)")

# ─────────────────────────────────────────────
# PROCESSAR PASTA
# ─────────────────────────────────────────────
def processar_pasta(pasta, saida="lancamentos.xlsx", banco_padrao=None):
    pasta=Path(pasta)
    if not pasta.exists(): print(f"❌ Pasta não encontrada: {pasta}"); return
    banco=banco_padrao or C["banco_itau"]; todos=[]
    xmls=sorted(pasta.glob("*.[xX][mM][lL]"))
    ofxs=sorted(pasta.glob("*.[oO][fF][xX]"))
    pdfs=sorted(pasta.glob("*.[pP][dD][fF]"))
    print(f"\n📂 {pasta}  →  {len(xmls)} XML | {len(ofxs)} OFX | {len(pdfs)} PDF")
    for f in xmls:
        print(f"  📄 {f.name}")
        try:
            lcts=ler_xml(f); todos.extend(lcts); print(f"     {len(lcts)} lançamentos")
        except Exception as e: print(f"     ⚠️ {e}")
    for f in ofxs:
        print(f"  🏦 {f.name}")
        try:
            lcts=ler_ofx(f,banco); todos.extend(lcts); print(f"     {len(lcts)} lançamentos")
        except Exception as e: print(f"     ⚠️ {e}")
    for f in pdfs:
        print(f"  📑 {f.name}")
        try:
            lcts=ler_pdf(f,banco); todos.extend(lcts); print(f"     {len(lcts)} lançamentos")
        except Exception as e: print(f"     ⚠️ {e}")
    if not todos: print("\n⚠️  Nenhum lançamento gerado."); return
    todos.sort(key=lambda x:x.data)
    gerar_excel(todos, saida)
    return todos

# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
if __name__=="__main__":
    ap=argparse.ArgumentParser(description="Contabilizador Automático")
    sub=ap.add_subparsers(dest="cmd")
    for nm,hlp in [("pasta","Processa pasta completa"),("xml","NF-e XML"),
                   ("ofx","Extrato OFX"),("pdf","PDF (DAS/boleto/folha/NF)"),("exemplo","Exemplo manual")]:
        p=sub.add_parser(nm,help=hlp)
        if nm!="exemplo":
            if nm=="pasta": p.add_argument("caminho")
            else: p.add_argument("arquivo")
            p.add_argument("-o","--saida",default="lancamentos.xlsx")
        if nm in ("pasta","ofx","pdf"):
            p.add_argument("--banco",default="itau",choices=list(BANCOS_MAP.keys()))
    args=ap.parse_args()
    def gb(n): return C.get(BANCOS_MAP.get(n,"banco_itau"),C["banco_itau"])
    if args.cmd=="pasta": processar_pasta(args.caminho,args.saida,gb(args.banco))
    elif args.cmd=="xml":
        lcts=ler_xml(args.arquivo)
        if lcts: gerar_excel(lcts,args.saida)
    elif args.cmd=="ofx":
        lcts=ler_ofx(args.arquivo,gb(args.banco))
        if lcts: gerar_excel(lcts,args.saida)
    elif args.cmd=="pdf":
        lcts=ler_pdf(args.arquivo,gb(args.banco))
        if lcts: gerar_excel(lcts,args.saida)
    elif args.cmd=="exemplo":
        lcts=[]
        lcts.extend(lancar_prolabore("31/01/2025",3000.00,"Sócio Admin"))
        lcts.extend(lancar_salario("31/01/2025","João Silva",2000.00,165.00,160.00))
        lcts.extend(lancar_das("20/01/2025",450.00,"01/2025"))
        gerar_excel(lcts,"exemplo_lancamentos.xlsx")
    else: ap.print_help()

